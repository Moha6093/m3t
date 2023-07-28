import openpyxl

from pywebio import start_server
from pywebio.input import *
from pywebio.output import *
from pywebio.session import *
from pywebio.pin import *

def App():
    put_html('<center><h2>Welcome to MT English Academy</h2></center>').style('background-color:#FFD95A; color:white; padding:10px;')
    put_html('<p>مرحبا بشباب المستقبل</p>').style('text-align:center; font-weight:bold')
    put_html('<center><img src="https://img.freepik.com/premium-vector/success-student-consulting_7109-29.jpg"></img></center>')

    SeS = input_group(
        'املأ الاستمارة الآتية',
        [
            input('اسم الطالب', name='student'),
            input('رقم الطالب', name='phone_num1', type=NUMBER),
            input('رقم ولي أمر الطالب', name='phone_num2', type=NUMBER),
            radio('الصف الدراسي', options=['First', 'Second', 'Third'], name='Grade'),
        ],
    )

    # Open the existing Excel file
    workbook = openpyxl.load_workbook('student_data.xlsx')
    sheet = workbook.active

    # Add the new data to the Excel file
    row = sheet.max_row + 1
    sheet.cell(row, 1).value = SeS['student']
    sheet.cell(row, 2).value = SeS['phone_num1']
    sheet.cell(row, 3).value = SeS['phone_num2']
    sheet.cell(row, 4).value = SeS['Grade']

    # Save the Excel file
    workbook.save('student_data.xlsx')

    # Display the student data in a table
    data = [
        ['Name', 'Student Num', 'Parent Num', 'Grade'],
        [SeS['student'], SeS['phone_num1'], SeS['phone_num2'], SeS['Grade']]
    ]
    put_table(data)

start_server(App, port=5000, debug=True)